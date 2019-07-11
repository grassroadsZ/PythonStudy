'''
-*-conding:utf-8
@Time:2019-05-21 6:40
@auther:grassroadsZ
@file:handle_excel.py
'''
from openpyxl import load_workbook
from collections import namedtuple
from Homework.lesson_21_0520_fina_rewrite.handle_config import do_config

class ExcelOption(object):
    """
    excel 操作类
    """
    def __init__(self,filename,sheetname=None):
        self.filename = filename
        # self.filename = do_config("file path","excel_path")
        # print(self.filename)
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active if self.sheetname is None else self.wb[self.sheetname]
        self.case_list = []
        self.Case = namedtuple( "Case" , tuple( self.ws.iter_rows( max_row = 1 , values_only = True ) )[0] )

    def get_cases(self):
        """
        获取所有的测试用例
        :return:所有的测试用例
        """
        Case = tuple( self.ws.iter_rows( min_row = 2 , values_only = True ) )
        for data in Case:
            self.case_list.append( self.Case( *data ) )
        return self.case_list

    def get_case(self,row):
        """

        :param row:用例行号
        :return: 行号对应的测试用例
        """
        if isinstance(row,int) and (2 <= row <= self.ws.max_row):
            case = tuple( self.ws.iter_rows( min_row = row , max_row = row , values_only = True ) )[0]
            return self.Case( *case )
        else:
            print("只能是正整数")

    def excel_write(self,row,real_result,result):
        """

        :param row:写入的行号
        :param real_result: 测试实际结果
        :param result:测试结果 pass/fail
        :return:
        """
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[self.sheetname]
        if isinstance(row,int) and (2 <= row <= self.ws.max_row):
            self.ws.cell(row = row,column = do_config("column","act_column"),value = real_result)
            self.ws.cell(row = row,column = do_config("column","res_column"),value = result)
            self.wb.save(self.filename)
        else:
            print("传入行号有误，行号应大于1的整数")
do_excel = ExcelOption(do_config("file path","excel_path"))


