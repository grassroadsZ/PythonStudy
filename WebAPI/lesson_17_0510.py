'''
-*-conding:utf-8
@Time:2019-05-13 7:13
@auther:grassroadsZ
@file:lesson_17_0510.py
'''

# 一、必做题
# 1.什么是ddt？它有什么作用？
# ddt 是一种数据驱动思想，也是一个辅助加载测试数据的数据库。

# 2.@unpack装饰器的作用？
# @unpack装饰器的作用是解决数据中需要多次拆包的情况。如嵌套字典的列表，其中字典的键同样是序列类型。

# 3.@data装饰器的使用规则？
# 先从ddt中导入，使用时与ddt库中的类装饰器ddt与类方法的装饰器data一起使用。

# 4.将之前写的两数相除的单元测试，使用ddt进行第一次重构

import unittest
import inspect
import time
from openpyxl import load_workbook
from ddt import ddt,data
from collections import namedtuple


# 打开excel
wb = load_workbook(r"F:\Python3.6\LemonPython_Study\Homework\testcase.xlsx")

#定位表单
ws = wb.active

case_list = []
# 获取表头
sheet_header = namedtuple("sheet_header",tuple(ws.iter_rows(max_row = 1,values_only = True))[0])
# 将每个用例组成字典放到list
data1 = tuple(ws.iter_rows(min_row = 2,values_only = True))
for data2 in data1:
    case_list.append(sheet_header(*data2))

class MathNum:
    def __init__(self,a ,b):
        self.a ,self.b = a, b

    def two_division(self):
        return self.a / self.b

    def two_sub(self):
        return self.a - self.b

@ddt
class TestDivision( unittest.TestCase ):
    """测试两数相除的类"""
    now = time.strftime( "%Y-%m-%d %H:%M:%S" , time.localtime() )
    #
    @classmethod
    def setUpClass(cls):
        cls.file_name = "test_result.txt"
        cls.file = open( cls.file_name , mode = "a" , encoding = "utf-8" )
        cls.file.write( "{:*^40s}".format( "测试开始" ) )

    @classmethod
    def tearDownClass(cls):
        cls.file.write( "{:*^40s}".format( "测试结束" ) )
        cls.file.close()
        wb.save(r"F:\Python3.6\LemonPython_Study\Homework\testcase.xlsx")

    @data(*case_list)
    def test_two_positive_division(self,case_list):
        """
        两正数相除
        :return: 当前测试用例的名称及测试结果
        """

        result_address = "测试时间：{}测试用例名称是{},测试结果是： ".format( TestDivision.now , inspect.stack()[0][3] )

        msg = case_list.tittle
        except_result = case_list.except_result
        first_num = case_list.first_num
        second_num = case_list.second_num
        real_result = case_list.real_result
        case_num = case_list.case_num

        try:
            real_result = MathNum( first_num , second_num ).two_division()
            self.assertEqual( except_result , real_result , msg = msg )
        except AssertionError as a:
            self.file.write("\n{}不通过,原因是{}\n".format(result_address,a))
            raise a
            ws.cell( row = case_num + 1 , column = 7 , value = "Fail" )
        else:
            self.file.write( "\n{},Pass\n".format(result_address + msg) )
            ws.cell(row = case_num+1,column =6 ,value = real_result)
            ws.cell( row = case_num + 1 , column = 7 , value = "Pass" )

if __name__ == '__main__':
    unittest.main()