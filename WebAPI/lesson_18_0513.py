'''
-*-conding:utf-8
@Time:2019-05-14 21:27
@auther:grassroadsZ
@file:lesson_18_0513.py
'''


# 一、必做题
# 1.如何将一些操作（Excel读写）进行封装？（写出封装的思路）
# 将通用的操作写成方法
# 将写死的数据提取出来写成属性
# 将其它实例方法需要调用构造方法的变量定义成属性
# 2.配置文件由哪几部分构成？
# 配置文件由:  # 注释
#             [section]---区域片段
#             option:value    --选项区

# 3.使用ConfigParser操作配置文件的步骤？
# from configparser import ConfigParser
# 创建配置文件对象
# config = ConfigParser()
# 指定读取的配置文件名
# read_file = config.read(filenames,encoding = "utf-8")
# 读取配置信息
# config.sections()
# config[section][option]


# 4.完成如下程序
#
# a.将Excel操作进行封装
#
# b.Excel操作封装之后，将之前写的两数相除的单元测试再次进行重构



# 5.完成如下程序
#
# a.将两数相除单元测试中的某些固定的参数抽离出来，放到配置文件中
#
# b.通过Python中的configparser模块，来读取配置文件中的数据
import unittest,inspect,time
from openpyxl import load_workbook
from ddt import ddt,data
from collections import namedtuple
from configparser import ConfigParser

class MathNum:
    def __init__(self,a ,b):
        self.a ,self.b = a, b

    def two_division(self):
        return self.a / self.b

class GetConfig(object):
    """
    获取配置文件类
    """
    def __init__(self):
        self.filename = r"F:\Python3.6\LemonPython_Study\Homework\case_config.conf"
        # 构建配置文件对象
        self.config = ConfigParser()
        # 指定读取的配置文件
        self.read_config = self.config.read(self.filename,encoding = "utf-8")
        # 读取数据
        self.config.sections()


    def get_config(self,sectionname,optionname):
        return self.config.get(sectionname,optionname)

class ExcelOption(object):
    """
    excel 操作类
    """
    # filename = r"F:\Python3.6\LemonPython_Study\Homework\case_config.conf"
    def __init__(self,filename,sheetname=None):
        self.filename = filename
        # self.filename = GetConfig().get_config("file path","excel_path")
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active if self.sheetname is None else self.wb[self.sheetname]
        self.case_list = []
        self.Case = namedtuple( "Case" , tuple( self.ws.iter_rows( max_row = 1 , values_only = True ) )[0] )

    def get_cases(self):
        """
        获取所有的测试用例
        :return:所有的测试用例
        """
        Case = tuple( self.ws.iter_rows( min_row = 2 , values_only = True ) )
        for data in Case:
            self.case_list.append( self.Case( *data ) )
        return self.case_list

    def get_case(self,row):
        """

        :param row:用例行号
        :return: 行号对应的测试用例
        """
        if isinstance(row,int) and (2 <= row <= self.ws.max_row):
            case = tuple( self.ws.iter_rows( min_row = row , max_row = row , values_only = True ) )[0]
            return self.Case( *case )
        else:
            print("只能是正整数")

    def excel_write(self,row,real_result,result):
        """

        :param row:写入的行号
        :param real_result: 测试实际结果
        :param result:测试结果 pass/fail
        :return:
        """
        self.ws.cell(row = row,column = int(GetConfig().get_config("column","act_column")),value = real_result)
        self.ws.cell(row = row,column = int(GetConfig().get_config("column","res_column")),value = result)
        self.wb.save(self.filename)





@ddt
class TestDivision( unittest.TestCase ):
    """测试两数相除的类"""
    now = time.strftime( "%Y-%m-%d %H:%M:%S" , time.localtime() )
    exl = ExcelOption(filename = GetConfig().get_config("file path","excel_path"))
    case_list = exl.get_cases()

    @classmethod
    def setUpClass(cls):
        cls.file_name = GetConfig().get_config("file path","report_path")
        cls.file = open( cls.file_name , mode = "a" , encoding = "utf-8" )
        cls.file.write( "{:*^40s}".format( "测试开始" ) )

    @classmethod
    def tearDownClass(cls):
        cls.file.write( "{:*^40s}".format( "测试结束" ) )
        cls.file.close()



    @data(*case_list)
    def test_two_positive_division(self,case_list):
        """
        两正数相除
        :return: 当前测试用例的名称及测试结果
        """

        result_address = "测试时间：{}测试用例名称是{},测试结果是： ".format( TestDivision.now , inspect.stack()[0][3] )

        msg = case_list.tittle
        except_result = case_list.except_result
        first_num = case_list.first_num
        second_num = case_list.second_num
        act_result = case_list.act_result
        case_num = case_list.case_num

        try:
            act_result = MathNum( first_num , second_num ).two_division()
            self.assertEqual( except_result , act_result , msg = msg )
        except AssertionError as a:
            self.file.write("\n{}不通过,原因是{}\n".format(result_address,a))
            self.exl.excel_write( row = case_num + 1 , real_result = act_result , result = GetConfig().get_config("result","Fail") )
            raise a

        else:
            self.file.write( "\n{},Pass\n".format(result_address + msg) )
            self.exl.excel_write( row = case_num+1 , real_result = act_result , result = GetConfig().get_config("result","success") )





