'''
-*-conding:utf-8
@Time:2019/5/9 22:20
@auther:grassroadsZ
@file:lesson_16_0508.py
'''


# 一、必做题
# 1.使用openpyxl操作Excel的步骤
# 一.导入openpyxl模块中的load_workbook
# 二.实例化一个由xlsx文件构成的workbook对象(打开excel文件)
# 三.实例化一个由表单构成的worksheet对象(定位表单)
#     workbook.active/workbook["sheetname"]
# 四.实例化一个单元格对象或iter_rows()构造一个生成器(定位单元格)
# 对单元格进行操作后save

# 2.定位表单有哪些方法？
# 通过workbook["sheetname"]或是通过workbook.active方法获取第一个表单


# 3.获取Excel单元格的数据有哪些方法？
# 获取excel数据的方法：
# 通过worksheet对象的cell(row= ,colum= value_only=True)获取
# 通过遍历每个单元格for row in range(workbook_min_row+1,workbook_max_row+1):
#                         for colum in range(workbook.min_colum+1,workbook.max_colum+1)
#                                 date = worksheet.cell(row=row,colum=colum).value
# 通过iter_rows构造一个生成器
# 通过指定需要处理的单元格worksheet["开始单元格：结束单元格"]
# 4.Excel中读取的数据类型有哪些？
# 数字类型：int,float,bool,complex
# 其余为字符串
# None
# 5.eval函数的作用？使用时有哪些注意事项？
# 反序列化，将从文件中符合python语法格式的数据类型还原
# 对于用户输入的内容不需要用eval转换

# 6.完成如下程序
# a.使用Excel来记录两数相除的测试用例（手动写入数据）
# b.使用openpyxl将所有用例数据读取出来
# c.分别使用多种方法来封装测试用例数据（嵌套字典的列表和嵌套命名元组的列表）

from openpyxl import load_workbook

def excel_operate():
    """
    读取excel的数据
    :return:封装后的数据
    """
    # 打开excel
    wb = load_workbook(r"F:\Python3.6\LemonPython_Study\Homework\testcase.xlsx")

    #定位表单
    ws = wb.active

    case_list = []
    # 获取表头
    sheet_hader = tuple(ws.iter_rows(max_row = 1,values_only = True))[0]
    # 将每个用例组成字典放到list
    for data in tuple(ws.iter_rows(min_row = 2,values_only = True)):
        case_list.append(dict(zip(sheet_hader,data)))
    return case_list



# 命名元祖空间留待研究后再补充

if __name__ == '__main__':
    print(excel_operate())

