# <editor-fold desc="Description">
'''
-*-conding:utf-8
@Time:2019-05-16 21:13
@auther:grassroadsZ
@file:lesson_19_0515.py
'''
from configparser import ConfigParser
# 1.使用两种方式封装配置文件的相关操作

# 第一种参考上课讲的方式
# 第二种参考我在班级群里发的文件（lemon_05_parser_config_5.py）
class GetConfig(object):
    """
    获取配置文件类
    """
    def __init__(self):
        self.filename = r"F:\Python3.6\LemonPython_Study\Homework\case_config.conf"
        # 构建配置文件对象
        self.config = ConfigParser()
        # 指定读取的配置文件
        self.read_config = self.config.read(self.filename,encoding = "utf-8")
        # 读取数据
        self.config.sections()


    def get_num(self,sectionname,optionname):
        # 通过传具体的区域与选项获取具体的值
        data = self.config.get(sectionname,optionname)
        if data.isdigit():
            return int(data)
        try:
            return float(data)
        except Exception:
            pass

    def get_other(self,sectionname,optionname=None,is_eval=False,is_bool=False):
        # 根据传值返回不同类型的数据
        if optionname is None:
            return dict(self.config[sectionname])
        if isinstance(is_bool,bool):
            if is_bool:
                return self.config.getboolean(sectionname,optionname)
            else:
                return self.config.get(sectionname,optionname)
        else:
            raise ValueError("is_bool的值必须是布尔类型")
        if isinstance(is_eval,bool):
            if is_eval is True:
                return eval(self.config.get(sectionname,optionname))
            else:
                return self.config.get(sectionname,optionname)
        else:
            raise ValueError("is_bool的值必须是布尔类型")


class ExcelConfig(ConfigParser):
    def __init__(self):
        # 继承父类，同时重写父类self.filename属性
        super().__init__()
        self.filename = r"F:\Python3.6\LemonPython_Study\Homework\case_config.conf"
        self.read(self.filename,encoding = "utf-8")

    def __call__(self , section , option=None , is_eval=False , is_bool=False):
        if option is None:
            return dict( self[section] )

        if isinstance(is_bool,bool):
            if is_bool:
                return self.getboolean( section , option )
            else:
                return self.get( section , option )
        else:
            raise ValueError("is_bool的值必须是布尔类型")

        if isinstance(is_eval,bool):
            if is_eval is True:
                return eval( self.get( section , option ) )
            else:
                return self.get( section , option )
        else:
            raise ValueError("is_bool的值必须是布尔类型")

        data = self.get( section , option )
        if data.isdigit():
            return int( data )
        try:
            return float( data )
        except Exception:
            pass


# 2.完成如下程序
# a.将两数相除单元测试中的某些固定的参数抽离出来，放到配置文件中
# b.通过第1题中自己封装的模块，来读取配置文件中的数据
# 提示：
    # a为上一次作业中的需求，可直接复制粘贴过来，只完成b条件即可。

import unittest,inspect,time
from openpyxl import load_workbook
from ddt import ddt,data
from collections import namedtuple
from configparser import ConfigParser
import HTMLTestRunnerNew

class MathNum:
    def __init__(self,a ,b):
        self.a ,self.b = a, b

    def two_division(self):
        return self.a / self.b
class ExcelOption(object):
    """
    excel 操作类
    """
    # filename = r"F:\Python3.6\LemonPython_Study\Homework\case_config.conf"
    def __init__(self,filename,sheetname=None):
        self.filename = filename
        # self.filename = GetConfig().get_config("file path","excel_path")
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active if self.sheetname is None else self.wb[self.sheetname]
        self.case_list = []
        self.Case = namedtuple( "Case" , tuple( self.ws.iter_rows( max_row = 1 , values_only = True ) )[0] )

    def get_cases(self):
        """
        获取所有的测试用例
        :return:所有的测试用例
        """
        Case = tuple( self.ws.iter_rows( min_row = 2 , values_only = True ) )
        for data in Case:
            self.case_list.append( self.Case( *data ) )
        return self.case_list

    def get_case(self,row):
        """

        :param row:用例行号
        :return: 行号对应的测试用例
        """
        if isinstance(row,int) and (2 <= row <= self.ws.max_row):
            case = tuple( self.ws.iter_rows( min_row = row , max_row = row , values_only = True ) )[0]
            return self.Case( *case )
        else:
            print("只能是正整数")

    def excel_write(self,row,real_result,result):
        """

        :param row:写入的行号
        :param real_result: 测试实际结果
        :param result:测试结果 pass/fail
        :return:
        """
        self.ws.cell(row = row,column = GetConfig().get_num("column","act_column"),value = real_result)
        self.ws.cell(row = row,column = GetConfig().get_num("column","res_column"),value = result)
        self.wb.save(self.filename)





@ddt
class TestDivision( unittest.TestCase ):
    """测试两数相除的类"""
    now = time.strftime( "%Y-%m-%d %H:%M:%S" , time.localtime() )
    exl = ExcelOption(filename = GetConfig().get_other("file path","excel_path"))
    case_list = exl.get_cases()


    @classmethod
    def setUpClass(cls):
        cls.file_name = GetConfig().get_other("file path","report_path")
        cls.file = open( cls.file_name , mode = "a" , encoding = "utf-8" )
        cls.file.write( "{:*^40s}".format( "测试开始" ) )



    @classmethod
    def tearDownClass(cls):
        cls.file.write( "{:*^40s}".format( "测试结束" ) )
        cls.file.close()



    @data(*case_list)
    def test_two_positive_division(self,case_list):
        """
        两正数相除
        :return: 当前测试用例的名称及测试结果
        """

        result_address = "测试时间：{}测试用例名称是{},测试结果是： ".format( TestDivision.now , inspect.stack()[0][3] )

        msg = case_list.tittle
        except_result = case_list.except_result
        first_num = case_list.first_num
        second_num = case_list.second_num
        act_result = case_list.act_result
        case_num = case_list.case_num

        try:
            act_result = MathNum( first_num , second_num ).two_division()
            self.assertEqual( except_result , act_result , msg = msg )
        except AssertionError as a:
            self.file.write("\n{}不通过,原因是{}\n".format(result_address,a))
            self.exl.excel_write( row = case_num + 1 , real_result = act_result , result = GetConfig().get_other("result","Fail") )
            raise a

        else:
            self.file.write( "\n{},Pass\n".format(result_address + msg) )
            self.exl.excel_write( row = case_num+1 , real_result = act_result , result = GetConfig().get_other("result","success") )

if __name__ == '__main__':
    # print(GetConfig().get_other("column",optionname = "res_column",is_bool =False))
    # unittest.main()
    load = unittest.TestLoader()
    one_suite = load.loadTestsFromTestCase( TestDivision )
    # load = unittest.TestSuite()
    # one_suite = load.addTests(TestDivision().test_two_positive_division)

    with open( "report.html" , "wb" ) as file:
        test_runner = HTMLTestRunnerNew.HTMLTestRunner( stream = file , verbosity = 2 , title = "测试报告" ,
                                                        description = "两数相加相除" , tester = "grassroadsZ" )
        test_runner.run( one_suite )

